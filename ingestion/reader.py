from pathlib import Path
from openpyxl import load_workbook


REQUIRED_SHEETS = [
    "Zerto Data",
    "Hypervisor Data",
    "Default VPG Settings",
    "Recovery ZVM Sites",
]


def load_excel_workbook(file_path: str):
    """
    Load an Excel workbook and return the workbook object.
    """
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    if path.suffix.lower() not in [".xlsx", ".xlsm"]:
        raise ValueError("Only .xlsx and .xlsm Excel files are supported")

    return load_workbook(filename=path, data_only=True)


def validate_required_sheets(workbook) -> None:
    """
    Check that all required sheets exist in the workbook.
    """
    existing_sheets = workbook.sheetnames

    missing_sheets = [
        sheet for sheet in REQUIRED_SHEETS
        if sheet not in existing_sheets
    ]

    if missing_sheets:
        raise ValueError(
            f"Missing required sheet(s): {', '.join(missing_sheets)}"
        )


def get_sheet(workbook, sheet_name: str):
    """
    Return a worksheet by name.
    """
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")

    return workbook[sheet_name]
